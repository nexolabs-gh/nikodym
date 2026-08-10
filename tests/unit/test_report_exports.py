"""Tests de ``report.exports``: las tablas por observación, completas y fuera del documento.

El criterio de aceptación es que el adjunto sirva **como dato**: completo (sin truncar), con el
identificador de la operación intacto y abrible por quien lo reciba. Los tests que escriben
``.xlsx`` van gateados con ``skipif`` sobre ``openpyxl`` (extra ``excel``), como en
``test_data_loading``: el job mínimo del CI instala sin extras y correrlos allí lo reventaría.
"""

from __future__ import annotations

import builtins
import importlib.util
import re
import sys
import tomllib
import warnings
from pathlib import Path

import pandas as pd
import pytest

from nikodym.report.config import ReportConfig, SectionPolicyConfig, XlsxExportConfig
from nikodym.report.document import PER_OBSERVATION_TABLES
from nikodym.report.exceptions import ReportDependencyError, ReportExportError
from nikodym.report.exports import (
    _XLSX_MISSING_MSG,
    DATA_EXPORT_FORMATS,
    _openpyxl_disponible,
    data_export_refs,
    per_observation_tables,
    write_data_exports,
)

_HAS_OPENPYXL = importlib.util.find_spec("openpyxl") is not None


def _score_frame(rows: int = 350) -> pd.DataFrame:
    """Frame por observación de ``rows`` filas, con ``loan_id`` como índice (el identificador)."""
    return pd.DataFrame(
        {
            "score": [600 + index for index in range(rows)],
            "pd_calibrated": [0.01 + index / 10_000 for index in range(rows)],
            "partición": ["desarrollo"] * rows,  # acento: el CSV debe conservarlo legible
        },
        index=pd.Index([f"op-{index:06d}" for index in range(rows)], name="loan_id"),
    )


def _tables() -> dict[str, pd.DataFrame]:
    """Bundle con una tabla por observación y una agregada (que NO debe exportarse)."""
    return {
        "scorecard.score": _score_frame(),
        "model.coefficients": pd.DataFrame({"feature": ["mora"], "beta": [1.25]}),
    }


def test_solo_las_tablas_por_observacion_son_adjuntos() -> None:
    """El criterio es la NATURALEZA de la tabla, no su tamaño: una agregada no sale del informe."""
    assert per_observation_tables(_tables()) == ("scorecard.score",)
    assert "model.coefficients" not in PER_OBSERVATION_TABLES
    assert frozenset({"csv", "xlsx"}) == DATA_EXPORT_FORMATS


def test_sin_csv_ni_xlsx_no_hay_adjuntos_ni_archivos(tmp_path: Path) -> None:
    """Sin formato de datos pedido no se escribe nada: el documento lo declara, no lo inventa."""
    config = ReportConfig(formats=())

    assert data_export_refs(_tables(), config=config) == ()
    assert write_data_exports(_tables(), config=config, output_dir=str(tmp_path)) == {}
    assert list(tmp_path.iterdir()) == []


def test_csv_completo_sin_truncar_y_con_el_identificador(tmp_path: Path) -> None:
    """El adjunto se escribe COMPLETO: ``max_table_rows`` acota lo que se muestra, no lo que se da.

    Es el punto entero del cambio: truncada a 200 filas, la tabla no servía ni como dato (estaba
    incompleta) ni como informe. Aquí se exige lo contrario: 350 filas dentro, 350 filas fuera.
    """
    config = ReportConfig(
        formats=("csv",),
        sections=SectionPolicyConfig(max_table_rows=10),  # el documento trunca; el dato NO
    )

    exports = write_data_exports(_tables(), config=config, output_dir=str(tmp_path))

    assert set(exports) == {"scorecard_report__scorecard_score.csv"}
    path = Path(exports["scorecard_report__scorecard_score.csv"])
    assert path.is_file()

    leido = pd.read_csv(path, encoding="utf-8-sig")
    assert len(leido) == 350  # completo, pese a max_table_rows=10
    assert leido.columns[0] == "loan_id"  # el índice es el identificador: no se pierde
    assert leido.iloc[0]["loan_id"] == "op-000000"
    assert leido.iloc[-1]["loan_id"] == "op-000349"
    assert "partición" in leido.columns  # el acento sobrevive al round-trip

    # La tabla agregada NO se exporta: se queda en el documento, que es donde se revisa.
    assert not list(tmp_path.glob("*coefficients*"))


def test_csv_es_byte_determinista(tmp_path: Path) -> None:
    """Dos escrituras del mismo frame producen bytes idénticos (reproducibilidad regulatoria)."""
    config = ReportConfig(formats=("csv",))

    primero = write_data_exports(_tables(), config=config, output_dir=str(tmp_path / "a"))
    segundo = write_data_exports(_tables(), config=config, output_dir=str(tmp_path / "b"))

    assert (
        Path(next(iter(primero.values()))).read_bytes()
        == Path(next(iter(segundo.values()))).read_bytes()
    )


def test_referencias_de_adjunto_son_puras_y_nombran_el_archivo_real(tmp_path: Path) -> None:
    """``data_export_refs`` (sin tocar disco) nombra exactamente los archivos que se escribirán.

    Es lo que permite que el documento diga "el detalle va en el archivo X" en el mismo render en
    que se decide emitirlo, sin adivinar el nombre ni referenciar un archivo inexistente.
    """
    config = ReportConfig(formats=("csv",))

    refs = data_export_refs(_tables(), config=config)
    exports = write_data_exports(_tables(), config=config, output_dir=str(tmp_path))

    assert [ref.filename for ref in refs] == list(exports)
    assert refs[0].table_key == "scorecard.score"
    assert refs[0].title == "Puntaje por observación"  # título legible, no la clave interna
    assert refs[0].rows == 350  # el tamaño REAL, no el truncado
    assert refs[0].sheet == ""  # el csv es un archivo por tabla, sin hojas


def test_export_crea_el_directorio_de_salida_si_falta(tmp_path: Path) -> None:
    """El export crea su directorio, igual que el HTML, el PDF y el ``.docx``: sin preflight."""
    destino = tmp_path / "aun" / "no" / "existe"

    exports = write_data_exports(
        _tables(), config=ReportConfig(formats=("csv",)), output_dir=str(destino)
    )

    assert Path(exports["scorecard_report__scorecard_score.csv"]).is_file()


def test_export_falla_con_error_accionable_si_la_ruta_no_es_escribible(tmp_path: Path) -> None:
    """Un fallo de escritura es un ``ReportExportError`` con acción, no un ``OSError`` desnudo."""
    ocupado = tmp_path / "ocupado"
    ocupado.write_text("no soy un directorio", encoding="utf-8")

    with pytest.raises(ReportExportError, match="acción="):
        write_data_exports(
            _tables(), config=ReportConfig(formats=("csv",)), output_dir=str(ocupado)
        )


@pytest.mark.skipif(not _HAS_OPENPYXL, reason="requiere el extra excel (openpyxl)")
def test_xlsx_un_libro_con_una_hoja_por_tabla(tmp_path: Path) -> None:
    """El ``.xlsx`` es UN libro con una hoja por tabla por observación, completa y con su índice."""
    import openpyxl

    config = ReportConfig(formats=("xlsx",), sections=SectionPolicyConfig(max_table_rows=10))
    tables = {**_tables(), "model.raw_pd_frame": _score_frame(rows=12)}

    exports = write_data_exports(tables, config=config, output_dir=str(tmp_path))

    assert set(exports) == {"scorecard_report__por_observacion.xlsx"}
    libro = openpyxl.load_workbook(exports["scorecard_report__por_observacion.xlsx"])
    assert libro.sheetnames == ["model_raw_pd_frame", "scorecard_score"]
    hoja = libro["scorecard_score"]
    assert hoja.max_row == 351  # 350 filas + encabezado: completo
    assert hoja.cell(row=1, column=1).value == "loan_id"
    assert hoja.cell(row=2, column=1).value == "op-000000"


@pytest.mark.skipif(not _HAS_OPENPYXL, reason="requiere el extra excel (openpyxl)")
def test_csv_y_xlsx_conviven_como_adjuntos_distintos(tmp_path: Path) -> None:
    """Pedir ambos formatos entrega ambos archivos; ninguno pisa al otro."""
    config = ReportConfig(formats=("csv", "xlsx"))

    exports = write_data_exports(_tables(), config=config, output_dir=str(tmp_path))

    assert set(exports) == {
        "scorecard_report__scorecard_score.csv",
        "scorecard_report__por_observacion.xlsx",
    }
    assert all(Path(path).is_file() for path in exports.values())


@pytest.mark.skipif(not _HAS_OPENPYXL, reason="requiere el extra excel (openpyxl)")
def test_nombre_de_hoja_excel_respeta_el_limite_de_31_caracteres() -> None:
    """Excel rechaza hojas de más de 31 caracteres: el nombre se recorta, no revienta al abrir.

    ⚠️ Gateado por el extra aunque no escriba nada: desde la degradación de M-5,
    ``data_export_refs`` **omite** las referencias del libro cuando ``openpyxl`` no está, así que
    sin el extra este test mediría el vacío en vez de los nombres de hoja.
    """
    config = ReportConfig(formats=("xlsx",))
    tables = {key: _score_frame(rows=2) for key in PER_OBSERVATION_TABLES}

    refs = data_export_refs(tables, config=config)

    assert len(refs) == len(PER_OBSERVATION_TABLES)
    hojas = [ref.sheet for ref in refs]
    assert all(len(hoja) <= 31 for hoja in hojas)
    assert len(set(hojas)) == len(hojas)  # sin colisiones tras el recorte


# ─────────────────────── M-5: la planilla degrada, no mata el informe ───────────────────────
#
# Hasta esta enmienda, `xlsx` era el único formato SIN degradación: `_write_xlsx` levantaba
# `ReportDependencyError` y nadie lo capturaba, así que la falta de un extra opcional no dejaba sin
# planilla — dejaba sin informe, incluido el HTML, que no depende de nada. Sus dos hermanos (`pdf`,
# `docx`) degradan en su renderer y tienen interruptor propio; aquí se cierra la asimetría.


class _SinOpenpyxl:
    """Finder que borra ``openpyxl`` del sistema de imports: ni ``find_spec`` ni ``import``.

    Simula la instalación **sin el extra** ``excel`` de forma determinista y por la ruta real, no
    parcheando la función que se quiere medir. Levantar ``ModuleNotFoundError`` (en vez de devolver
    ``None``) es lo que hace que ``importlib.util.find_spec`` también lo vea ausente.
    """

    def find_spec(self, name: str, path: object = None, target: object = None) -> None:
        if name == "openpyxl" or name.startswith("openpyxl."):
            raise ModuleNotFoundError(f"No module named {name!r}")
        return None


def _sin_extra_excel(monkeypatch: pytest.MonkeyPatch) -> None:
    """Deja el proceso como una instalación sin ``nikodym[excel]`` (ausencia real, no simulada)."""
    for modulo in [
        name for name in sys.modules if name == "openpyxl" or name.startswith("openpyxl.")
    ]:
        monkeypatch.delitem(sys.modules, modulo, raising=False)
    monkeypatch.setattr(sys, "meta_path", [_SinOpenpyxl(), *sys.meta_path])


def _openpyxl_presente_pero_roto(monkeypatch: pytest.MonkeyPatch) -> None:
    """Deja el paquete localizable pero su import roto: el caso de la SEGUNDA capa de defensa."""
    real_import = builtins.__import__

    def fake_import(
        name: str,
        globals_: dict[str, object] | None = None,
        locals_: dict[str, object] | None = None,
        fromlist: tuple[str, ...] = (),
        level: int = 0,
    ) -> object:
        if name == "openpyxl":
            raise ImportError("openpyxl instalado a medias")
        return real_import(name, globals_, locals_, fromlist, level)

    for modulo in [
        name for name in sys.modules if name == "openpyxl" or name.startswith("openpyxl.")
    ]:
        monkeypatch.delitem(sys.modules, modulo, raising=False)
    monkeypatch.setattr(builtins, "__import__", fake_import)


def test_sin_openpyxl_la_planilla_se_omite_y_el_resto_del_informe_sale(
    monkeypatch: pytest.MonkeyPatch, tmp_path: Path
) -> None:
    """CONTROL NEGATIVO de M-5: sin el extra, `xlsx` avisa y NO tumba el resto de los exports.

    Se piden los dos formatos de datos. Antes de la enmienda, el ``ReportDependencyError`` del libro
    propagaba desde ``write_data_exports`` y se llevaba por delante hasta el ``.csv`` ya escrito en
    disco, que quedaba fuera del resultado. Ahora el ``.csv`` sale y se devuelve.
    """
    _sin_extra_excel(monkeypatch)
    # Control positivo: el payload deja el extra realmente ausente, o sea lo que el caso mide.
    assert not _openpyxl_disponible()
    config = ReportConfig(formats=("csv", "xlsx"))

    with pytest.warns(RuntimeWarning, match=r"No se pudo generar el export \.xlsx"):
        exports = write_data_exports(_tables(), config=config, output_dir=str(tmp_path))

    assert set(exports) == {"scorecard_report__scorecard_score.csv"}
    assert Path(exports["scorecard_report__scorecard_score.csv"]).is_file()
    assert not list(tmp_path.glob("*.xlsx"))


def test_sin_openpyxl_el_documento_no_referencia_una_planilla_inexistente(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    """La referencia del libro se omite: el documento no puede nombrar un adjunto que no saldrá.

    ``data_export_refs`` es la fuente ÚNICA de nombres —la consultan el documento y el writer—, así
    que omitir ahí es lo que impide que las dos superficies divergan. El criterio ya estaba escrito
    en el docstring de la función: *«en vez de referenciar un archivo inexistente»*.
    """
    _sin_extra_excel(monkeypatch)
    config = ReportConfig(formats=("csv", "xlsx"))

    refs = data_export_refs(_tables(), config=config)

    assert [ref.filename for ref in refs] == ["scorecard_report__scorecard_score.csv"]
    assert all(ref.sheet == "" for ref in refs)


def test_sin_openpyxl_y_fail_if_unavailable_la_corrida_se_detiene(
    monkeypatch: pytest.MonkeyPatch, tmp_path: Path
) -> None:
    """El interruptor gemelo del ``.docx``: activado, la falta del extra detiene con diagnóstico."""
    _sin_extra_excel(monkeypatch)
    config = ReportConfig(formats=("csv", "xlsx"), xlsx=XlsxExportConfig(fail_if_unavailable=True))

    with pytest.raises(ReportDependencyError, match="nikodym\\[excel\\]"):
        write_data_exports(_tables(), config=config, output_dir=str(tmp_path))


@pytest.mark.skipif(not _HAS_OPENPYXL, reason="mide el extra PRESENTE pero roto: exige tenerlo")
def test_openpyxl_presente_pero_roto_tambien_degrada(
    monkeypatch: pytest.MonkeyPatch, tmp_path: Path
) -> None:
    """SEGUNDA capa: ``find_spec`` lo ve, el import falla, y el informe sigue saliendo.

    Es la razón por la que la captura de ``ReportDependencyError`` alrededor de ``_write_xlsx`` NO
    es código muerto tras añadir la comprobación previa: una instalación a medias o una versión
    incompatible con el ``ExcelWriter`` de pandas siguen levantando al importar de verdad.

    ⚠️ Gateado por el extra, y **el CI lo cazó donde el local no podía**: su control positivo
    afirma que ``openpyxl`` SÍ se localiza —es lo que distingue este caso del anterior— y en
    los jobs que instalan sin extras eso es falso. Un test que mide «presente pero roto»
    necesita que esté presente.
    """
    _openpyxl_presente_pero_roto(monkeypatch)
    # Control positivo: este caso NO es el anterior — el paquete se sigue localizando.
    assert _openpyxl_disponible()
    config = ReportConfig(formats=("csv", "xlsx"))

    with pytest.warns(RuntimeWarning, match=r"No se pudo generar el export \.xlsx"):
        exports = write_data_exports(_tables(), config=config, output_dir=str(tmp_path))

    assert set(exports) == {"scorecard_report__scorecard_score.csv"}
    assert not list(tmp_path.glob("*.xlsx"))


@pytest.mark.skipif(not _HAS_OPENPYXL, reason="mide el extra PRESENTE pero roto: exige tenerlo")
def test_openpyxl_presente_pero_roto_con_fail_if_unavailable_relanza(
    monkeypatch: pytest.MonkeyPatch, tmp_path: Path
) -> None:
    """El interruptor gobierna las DOS capas, no sólo la comprobación previa."""
    _openpyxl_presente_pero_roto(monkeypatch)
    config = ReportConfig(formats=("xlsx",), xlsx=XlsxExportConfig(fail_if_unavailable=True))

    with pytest.raises(ReportDependencyError, match="nikodym\\[excel\\]"):
        write_data_exports(_tables(), config=config, output_dir=str(tmp_path))


def test_sin_tablas_por_observacion_la_planilla_ausente_no_genera_ruido(
    monkeypatch: pytest.MonkeyPatch, tmp_path: Path
) -> None:
    """Sin nada que exportar no hay aviso: un aviso que se dispara de más se aprende a ignorar."""
    _sin_extra_excel(monkeypatch)
    config = ReportConfig(formats=("xlsx",))

    with warnings.catch_warnings():
        warnings.simplefilter("error")  # cualquier aviso aquí falla el test
        exports = write_data_exports(
            {"model.coefficients": pd.DataFrame({"beta": [1.0]})},
            config=config,
            output_dir=str(tmp_path),
        )

    assert exports == {}


def test_el_filtro_de_pyproject_ancla_el_mensaje_real_de_la_dependencia() -> None:
    """El ``filterwarnings`` del repo tiene que casar con el mensaje que el motor emite HOY.

    Sin este gate, cambiar el texto de ``_XLSX_MISSING_MSG`` deja el ``ignore`` sin aplicar **en
    silencio** y los jobs sin el extra caen con ``filterwarnings=["error"]``, lejos de la causa. El
    propio comentario del ``pyproject`` promete «cambiar ambos a la vez»; esto lo hace cumplir.
    """
    # Parseado con `tomllib`, no leído como texto: es la MISMA cadena que recibe pytest, ya sin los
    # escapes del TOML. Comparar contra el fuente crudo mediría el archivo, no el filtro efectivo.
    raiz = Path(__file__).resolve().parents[2]
    with (raiz / "pyproject.toml").open("rb") as archivo:
        filtros = tomllib.load(archivo)["tool"]["pytest"]["ini_options"]["filterwarnings"]
    patrones = [
        filtro.split(":")[1]
        for filtro in filtros
        if filtro.startswith("ignore:") and filtro.endswith(":RuntimeWarning")
    ]

    assert patrones, "no se encontró ningún filtro de RuntimeWarning en pyproject.toml"
    xlsx = [patron for patron in patrones if "xlsx" in patron]
    assert len(xlsx) == 1, f"se esperaba un único filtro de la planilla, hay {len(xlsx)}: {xlsx}"
    # `filterwarnings` casa el patrón contra el INICIO del mensaje (re.match), no contra todo.
    assert re.match(xlsx[0], _XLSX_MISSING_MSG), (
        f"el filtro {xlsx[0]!r} ya no casa con el mensaje real {_XLSX_MISSING_MSG!r}"
    )
